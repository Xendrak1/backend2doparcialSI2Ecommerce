from datetime import timedelta, date, datetime, time
import re
from django.db.models import Sum, Count, Avg
from django.db.models.functions import TruncDate
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from io import BytesIO
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
except Exception:
    A4 = None
    canvas = None
try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

from gestion.models import Venta, VentaDetalle, Stock, ProductoVariante, Producto


class ReporteResumen(APIView):
    """
    Resumen general para el dashboard de reportes.
    Devuelve totales de venta, ticket promedio, mix por canal y conteos clave.
    """
    def get(self, request):
        hoy = timezone.now().date()
        hace_30 = hoy - timedelta(days=30)

        # Solo contar ventas confirmadas y pagadas
        ventas_qs = Venta.objects.filter(estado='completado', estado_pago='pagado')
        ventas_30_qs = ventas_qs.filter(fecha__date__gte=hace_30)

        base_agg = ventas_qs.aggregate(total=Sum('total'), count=Count('id'))
        ventas_agg = {
            "total": base_agg.get("total") or 0,
            "count": base_agg.get("count") or 0,
            "promedio": 0,
        }
        if ventas_agg["count"] > 0:
            ventas_agg["promedio"] = float(ventas_agg["total"]) / float(ventas_agg["count"])
        ventas_30_agg = ventas_30_qs.aggregate(total=Sum('total'), count=Count('id'))

        # Mix por canal
        canales = list(
            ventas_qs.values('canal_venta')
            .annotate(total=Sum('total'), count=Count('id'))
            .order_by('-total')
        )

        # Mix por tipo de pago
        tipos_pago = list(
            ventas_qs.values('tipo_pago')
            .annotate(total=Sum('total'), count=Count('id'))
            .order_by('-total')
        )

        # Stock bajo (<= 5)
        stock_bajo = (
            Stock.objects.select_related('producto_variante__producto')
            .filter(cantidad__lte=5)
            .values(
                'producto_variante__producto__id',
                'producto_variante__producto__nombre',
            )
            .annotate(total_unidades=Sum('cantidad'))
            .order_by('total_unidades')[:10]
        )

        data = {
            "ventas": {
                "total": ventas_agg.get('total') or 0,
                "count": ventas_agg.get('count') or 0,
                "promedio": ventas_agg.get('promedio') or 0,
                "ultimos_30": {
                    "total": ventas_30_agg.get('total') or 0,
                    "count": ventas_30_agg.get('count') or 0,
                },
            },
            "canales": canales,
            "tipos_pago": tipos_pago,
            "stock_bajo": list(stock_bajo),
            "conteos": {
                "productos": Producto.objects.count(),
                "variantes": ProductoVariante.objects.count(),
            },
        }
        return Response(data)


class VentasPorDia(APIView):
    """
    Serie temporal de ventas por día en el rango indicado (default: últimos 30 días).
    Params opcionales: ?dias=30 o ?start=YYYY-MM-DD&end=YYYY-MM-DD
    """
    def get(self, request):
        hoy = timezone.now().date()
        desde = None
        hasta = None
        
        # Priorizar start/end sobre dias
        start_str = request.query_params.get('start')
        end_str = request.query_params.get('end')
        
        if start_str:
            try:
                desde = datetime.strptime(start_str, '%Y-%m-%d').date()
            except:
                pass
        if end_str:
            try:
                end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
                # Incluir el día completo de "end" (hasta el inicio del día siguiente)
                hasta = end_date + timedelta(days=1)
            except:
                pass
        
        # Si solo hay start, usar hasta hoy
        if desde and not hasta:
            hasta = hoy + timedelta(days=1)
        # Si solo hay end, usar desde el inicio de los tiempos (o desde hace mucho)
        if hasta and not desde:
            desde = hoy - timedelta(days=365)  # Último año por defecto
        
        # Si no hay start/end, usar dias
        if not desde and not hasta:
            dias = int(request.query_params.get('dias', 30))
            desde = hoy - timedelta(days=dias)
            hasta = hoy + timedelta(days=1)  # Incluir hoy
        
        # Construir el queryset - solo ventas confirmadas y pagadas
        qs = Venta.objects.filter(estado='completado', estado_pago='pagado')
        if desde:
            qs = qs.filter(fecha__date__gte=desde)
        if hasta:
            qs = qs.filter(fecha__date__lt=hasta)
        
        serie = (
            qs.annotate(dia=TruncDate('fecha'))
            .values('dia')
            .annotate(total=Sum('total'), count=Count('id'))
            .order_by('dia')
        )
        return Response(list(serie))


class TopProductos(APIView):
    """
    Top N productos por unidades o monto.
    Params: ?limit=5&metric=unidades|monto&order=desc|asc&start=YYYY-MM-DD&end=YYYY-MM-DD&season=otono|invierno|primavera|verano&year=YYYY&month=1-12&canal=tienda|online&categoria=<id>&exclude=nombre1,nombre2
    """
    def get(self, request):
        limit = int(request.query_params.get('limit', 5))
        metric = request.query_params.get('metric', 'unidades')
        order = (request.query_params.get('order') or 'desc').lower()
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        season = (request.query_params.get('season') or '').lower()
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        
        # Debug: imprimir parámetros recibidos
        min_precio_unitario = request.query_params.get('min_precio_unitario')
        max_precio_unitario = request.query_params.get('max_precio_unitario')
        min_monto = request.query_params.get('min_monto')
        max_monto = request.query_params.get('max_monto')
        print(f"[TopProductos] Params: limit={limit}, metric={metric}, order={order}, start={start}, end={end}, season={season}, year={year}, month={month}, min_precio_unitario={min_precio_unitario}, max_precio_unitario={max_precio_unitario}, min_monto={min_monto}, max_monto={max_monto}")
        
        # Solo contar detalles de ventas confirmadas y pagadas
        qs = VentaDetalle.objects.select_related('venta', 'producto_variante__producto').filter(
            venta__estado='completado',
            venta__estado_pago='pagado'
        )
        # Filtros de fecha
        if start:
            try:
                d = date.fromisoformat(start)
                dt_start = timezone.make_aware(datetime.combine(d, time.min))
                qs = qs.filter(venta__fecha__gte=dt_start)
            except Exception:
                pass
        if end:
            try:
                d = date.fromisoformat(end)
                # Fin de día inclusivo: < (end + 1 día) 00:00
                dt_end = timezone.make_aware(datetime.combine(d + timedelta(days=1), time.min))
                qs = qs.filter(venta__fecha__lt=dt_end)
            except Exception:
                pass
        if year:
            try:
                qs = qs.filter(venta__fecha__year=int(year))
            except Exception:
                pass
        if month:
            try:
                qs = qs.filter(venta__fecha__month=int(month))
            except Exception:
                pass
        if season in {'otono', 'otoño', 'invierno', 'primavera', 'verano'}:
            # Hemisferio sur: Otoño(3-5), Invierno(6-8), Primavera(9-11), Verano(12,1,2)
            season_map = {
                'otoño': [3, 4, 5],
                'otono': [3, 4, 5],
                'invierno': [6, 7, 8],
                'primavera': [9, 10, 11],
                'verano': [12, 1, 2],
            }
            months = season_map.get(season, [])
            if months:
                qs = qs.filter(venta__fecha__month__in=months)
        # Filtros adicionales
        canal = request.query_params.get('canal')
        if canal:
            qs = qs.filter(venta__canal_venta=canal)
        categoria = request.query_params.get('categoria')
        if categoria:
            try:
                qs = qs.filter(producto_variante__producto__categoria_id=int(categoria))
            except Exception:
                pass
        # Exclusiones por nombre de producto (icontains), CSV
        exclude = request.query_params.get('exclude')
        if exclude:
            # Construimos iregex tolerante a acentos y espacios
            def to_iregex_like(s: str) -> str:
                s = s.strip().lower()
                map_chars = {
                    'a': '[aá]',
                    'e': '[eé]',
                    'i': '[ií]',
                    'o': '[oó]',
                    'u': '[uú]',
                    'n': '[nñ]',
                    'A': '[AÁaá]',
                    'E': '[EÉeé]',
                    'I': '[IÍií]',
                    'O': '[OÓoó]',
                    'U': '[UÚuú]',
                    'N': '[NÑnñ]',
                }
                parts = []
                for ch in s:
                    if ch in map_chars:
                        parts.append(map_chars[ch])
                    elif ch.isalnum():
                        parts.append(re.escape(ch))
                    elif ch.isspace():
                        parts.append(r'\s+')
                    else:
                        parts.append(re.escape(ch))
                return ''.join(parts)
            for raw in exclude.split(','):
                name = (raw or '').strip()
                if name:
                    pattern = to_iregex_like(name)
                    qs = qs.exclude(producto_variante__producto__nombre__iregex=pattern)
        # Filtros por precio unitario (ANTES de values/annotate)
        # (ya obtenidos arriba en el debug)
        if min_precio_unitario:
            try:
                qs = qs.filter(producto_variante__precio__gte=float(min_precio_unitario))
            except Exception:
                pass
        if max_precio_unitario:
            try:
                qs = qs.filter(producto_variante__precio__lte=float(max_precio_unitario))
            except Exception:
                pass
        
        qs = qs.values('producto_variante__producto__id', 'producto_variante__producto__nombre')
        if metric == 'monto':
            qs = qs.annotate(valor=Sum('subtotal'))
        else:
            qs = qs.annotate(valor=Sum('cantidad'))
        
        # Filtros por monto mínimo/máximo (solo para metric=monto, DESPUÉS de annotate)
        # (ya obtenidos arriba en el debug)
        if min_monto and metric == 'monto':
            try:
                qs = qs.filter(valor__gte=float(min_monto))
            except Exception:
                pass
        if max_monto and metric == 'monto':
            try:
                qs = qs.filter(valor__lte=float(max_monto))
            except Exception:
                pass
        
        # Si piden "menos vendido", orden ascendente y omitimos cero para evitar ruido
        if order == 'asc':
            qs = qs.filter(valor__gt=0).order_by('valor', 'producto_variante__producto__nombre')[:limit]
        else:
            qs = qs.order_by('-valor', 'producto_variante__producto__nombre')[:limit]
        return Response(list(qs))


class MixPago(APIView):
    """
    Distribución por tipo de pago basado en Venta.tipo_pago.
    """
    def get(self, request):
        # Solo contar ventas confirmadas y pagadas
        mix = (
            Venta.objects.filter(estado='completado', estado_pago='pagado')
            .values('tipo_pago')
            .annotate(total=Sum('total'), count=Count('id'))
            .order_by('-total')
        )
        return Response(list(mix))


class StockBajo(APIView):
    """
    Lista de productos con stock por debajo o igual al umbral.
    Params: ?umbral=5&limit=20
    """
    def get(self, request):
        umbral = int(request.query_params.get('umbral', 5))
        limit = int(request.query_params.get('limit', 20))
        items = (
            Stock.objects.select_related('producto_variante__producto')
            .filter(cantidad__lte=umbral)
            .values(
                'producto_variante__producto__id',
                'producto_variante__producto__nombre',
            )
            .annotate(total_unidades=Sum('cantidad'))
            .order_by('total_unidades')[:limit]
        )
        return Response(list(items))


def _build_summary():
    hoy = timezone.now().date()
    hace_30 = hoy - timedelta(days=30)
    # Solo contar ventas confirmadas y pagadas
    ventas_qs = Venta.objects.filter(estado='completado', estado_pago='pagado')
    base_agg = ventas_qs.aggregate(total=Sum('total'), count=Count('id'))
    total = base_agg.get("total") or 0
    count = base_agg.get("count") or 0
    promedio = float(total) / float(count) if count else 0
    ventas_30 = ventas_qs.filter(fecha__date__gte=hace_30).aggregate(total=Sum('total'), count=Count('id'))
    canales = list(
        ventas_qs.values('canal_venta').annotate(total=Sum('total'), count=Count('id')).order_by('-total')
    )
    tipos_pago = list(
        ventas_qs.values('tipo_pago').annotate(total=Sum('total'), count=Count('id')).order_by('-total')
    )
    return {
        "total": total, "count": count, "promedio": promedio,
        "ultimos_30": {"total": ventas_30.get('total') or 0, "count": ventas_30.get('count') or 0},
        "canales": canales, "tipos_pago": tipos_pago,
    }


class ExportResumenPDF(APIView):
    def get(self, request):
        # Obtener parámetros de filtro
        recurso = request.query_params.get('recurso', 'resumen')
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        dias = request.query_params.get('dias')
        metric = request.query_params.get('metric', 'unidades')
        
        # Aplicar filtros de fecha si existen
        # Solo contar ventas confirmadas y pagadas
        qs_ventas = Venta.objects.filter(estado='completado', estado_pago='pagado')
        if start:
            try:
                d = date.fromisoformat(start)
                dt_start = timezone.make_aware(datetime.combine(d, time.min))
                qs_ventas = qs_ventas.filter(fecha__gte=dt_start)
            except Exception:
                pass
        if end:
            try:
                d = date.fromisoformat(end)
                dt_end = timezone.make_aware(datetime.combine(d + timedelta(days=1), time.min))
                qs_ventas = qs_ventas.filter(fecha__lt=dt_end)
            except Exception:
                pass
        # Construir datos usando queryset filtrado
        base_agg = qs_ventas.aggregate(total=Sum('total'), count=Count('id'))
        total = base_agg.get("total") or 0
        count = base_agg.get("count") or 0
        promedio = float(total) / float(count) if count else 0
        canales = list(
            qs_ventas.values('canal_venta').annotate(total=Sum('total'), count=Count('id')).order_by('-total')
        )
        tipos_pago = list(
            qs_ventas.values('tipo_pago').annotate(total=Sum('total'), count=Count('id')).order_by('-total')
        )
        data = {
            "total": total, "count": count, "promedio": promedio,
            "canales": canales, "tipos_pago": tipos_pago,
        }
        
        if not canvas:
            # Fallback: enviar TXT si reportlab no está instalado
            txt = f"TOTAL: {data['total']}\nCANTIDAD: {data['count']}\nPROMEDIO: {data['promedio']}\n\n"
            if start or end:
                txt += f"PERÍODO: {start or 'inicio'} a {end or 'hoy'}\n\n"
            txt += "MIX CANAL:\n" + "\n".join([f"- {c['canal_venta']}: {c['total']} ({c['count']})" for c in data["canales"]]) + "\n\n"
            txt += "MIX PAGO:\n" + "\n".join([f"- {t['tipo_pago']}: {t['total']} ({t['count']})" for t in data["tipos_pago"]]) + "\n\n"
            # Ventas por día
            dias_filtro = int(dias) if dias else 7
            desde = timezone.now() - timedelta(days=dias_filtro)
            serie_qs = qs_ventas.filter(fecha__gte=desde) if not start else qs_ventas
            serie = (
                serie_qs
                .annotate(dia=TruncDate('fecha'))
                .values('dia')
                .annotate(total=Sum('total'), count=Count('id'))
                .order_by('dia')
            )
            txt += f"VENTAS POR DÍA ({dias_filtro}d):\n" + "\n".join([f"- {s['dia']}: {s['total']} ({s['count']})" for s in serie])
            resp = HttpResponse(txt.encode("utf-8"), content_type="text/plain")
            nombre_archivo = f"reporte_{recurso}"
            if start and end:
                nombre_archivo += f"_{start}_{end}"
            resp["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.txt"'
            return resp
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        p.setTitle(f"Reporte {recurso}")
        p.setFont("Helvetica-Bold", 16)
        titulo = f"Reporte - {recurso.replace('_', ' ').title()}"
        if start or end:
            titulo += f" ({start or 'inicio'} a {end or 'hoy'})"
        p.drawString(50, 800, titulo)
        p.setFont("Helvetica", 12)
        y = 770
        p.drawString(50, y, f"Total ventas: Bs. {data['total']:.2f}") ; y -= 18
        p.drawString(50, y, f"Cantidad: {data['count']}") ; y -= 18
        p.drawString(50, y, f"Promedio: Bs. {data['promedio']:.2f}") ; y -= 24
        p.drawString(50, y, "Mix por canal:") ; y -= 16
        for c in data["canales"]:
            p.drawString(60, y, f"- {c['canal_venta']}: Bs. {float(c['total'] or 0):.2f} ({c['count']})") ; y -= 16
        y -= 8
        p.drawString(50, y, "Mix por tipo de pago:") ; y -= 16
        for t in data["tipos_pago"]:
            p.drawString(60, y, f"- {t['tipo_pago']}: Bs. {float(t['total'] or 0):.2f} ({t['count']})") ; y -= 16
        y -= 24
        # Serie de días
        dias_filtro = int(dias) if dias else 7
        desde = timezone.now() - timedelta(days=dias_filtro)
        serie_qs = qs_ventas.filter(fecha__gte=desde) if not start else qs_ventas
        serie = (
            serie_qs
            .annotate(dia=TruncDate('fecha'))
            .values('dia')
            .annotate(total=Sum('total'), count=Count('id'))
            .order_by('dia')
        )
        p.drawString(50, y, f"Ventas por día ({dias_filtro}d):") ; y -= 16
        for s in serie:
            p.drawString(60, y, f"- {s['dia']}: Bs. {float(s['total'] or 0):.2f} ({s['count']})") ; y -= 16
        y -= 24
        # Top Productos (usando queryset filtrado)
        qs_detalles = VentaDetalle.objects.select_related('venta', 'producto_variante__producto').filter(venta__in=qs_ventas)
        if metric == 'monto':
            top = (
                qs_detalles
                .values('producto_variante__producto__nombre')
                .annotate(valor=Sum('subtotal'))
                .order_by('-valor')[:5]
            )
            p.drawString(50, y, "Top 5 productos (monto):") ; y -= 16
            for t in top:
                p.drawString(60, y, f"- {t['producto_variante__producto__nombre']}: Bs. {float(t['valor'] or 0):.2f}") ; y -= 16
        else:
            top = (
                qs_detalles
                .values('producto_variante__producto__nombre')
                .annotate(unidades=Sum('cantidad'))
                .order_by('-unidades')[:5]
            )
            p.drawString(50, y, "Top 5 productos (unidades):") ; y -= 16
            for t in top:
                p.drawString(60, y, f"- {t['producto_variante__producto__nombre']}: {t['unidades']} u.") ; y -= 16
        y -= 24
        # Stock Bajo
        stock_bajo = (
            Stock.objects.select_related('producto_variante__producto')
            .filter(cantidad__lte=5)
            .values('producto_variante__producto__nombre')
            .annotate(total_unidades=Sum('cantidad'))
            .order_by('total_unidades')[:10]
        )
        p.drawString(50, y, "Stock bajo (<=5):") ; y -= 16
        for sb in stock_bajo:
            p.drawString(60, y, f"- {sb['producto_variante__producto__nombre']}: {sb['total_unidades']} u.") ; y -= 16
        p.showPage()
        p.save()
        buffer.seek(0)
        resp = HttpResponse(buffer.read(), content_type="application/pdf")
        nombre_archivo = f"reporte_{recurso}"
        if start and end:
            nombre_archivo += f"_{start}_{end}"
        resp["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.pdf"'
        return resp


class ExportResumenExcel(APIView):
    def get(self, request):
        if not Workbook:
            # Fallback: CSV si openpyxl no está instalado
            import csv, io
            sio_text = io.StringIO()
            writer = csv.writer(sio_text, lineterminator="\n")
            data = _build_summary()
            writer.writerow(["Métrica", "Valor"])
            writer.writerow(["Total ventas", data["total"]])
            writer.writerow(["Cantidad", data["count"]])
            writer.writerow(["Promedio", data["promedio"]])
            csv_content = sio_text.getvalue()
            resp = HttpResponse(csv_content, content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="reporte_resumen.csv"'
            return resp
        data = _build_summary()
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"
        ws.append(["Métrica", "Valor"])
        ws.append(["Total ventas", data["total"]])
        ws.append(["Cantidad", data["count"]])
        ws.append(["Promedio", data["promedio"]])
        # Ventas por día (7d)
        desde = timezone.now() - timedelta(days=7)
        # Solo contar ventas confirmadas y pagadas
        serie = (
            Venta.objects.filter(estado='completado', estado_pago='pagado', fecha__gte=desde)
            .annotate(dia=TruncDate('fecha'))
            .values('dia')
            .annotate(total=Sum('total'), count=Count('id'))
            .order_by('dia')
        )
        ws4 = wb.create_sheet("VentasPorDia_7d")
        ws4.append(["Día", "Total", "Cantidad"])
        for s in serie:
            ws4.append([str(s["dia"]), s["total"], s["count"]])
        ws2 = wb.create_sheet("Canales")
        ws2.append(["Canal", "Total", "Cantidad"])
        for c in data["canales"]:
            ws2.append([c["canal_venta"], c["total"], c["count"]])
        ws3 = wb.create_sheet("TiposPago")
        ws3.append(["Tipo", "Total", "Cantidad"])
        for t in data["tipos_pago"]:
            ws3.append([t["tipo_pago"], t["total"], t["count"]])
        # Top productos
        top = (
            VentaDetalle.objects.select_related('producto_variante__producto')
            .values('producto_variante__producto__nombre')
            .annotate(unidades=Sum('cantidad'))
            .order_by('-unidades')[:5]
        )
        ws5 = wb.create_sheet("TopProductos")
        ws5.append(["Producto", "Unidades"])
        for t in top:
            ws5.append([t["producto_variante__producto__nombre"], t["unidades"]])
        # Stock bajo
        stock_bajo = (
            Stock.objects.select_related('producto_variante__producto')
            .filter(cantidad__lte=5)
            .values('producto_variante__producto__nombre')
            .annotate(total_unidades=Sum('cantidad'))
            .order_by('total_unidades')[:10]
        )
        ws6 = wb.create_sheet("StockBajo")
        ws6.append(["Producto", "Unidades"])
        for sb in stock_bajo:
            ws6.append([sb["producto_variante__producto__nombre"], sb["total_unidades"]])
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        resp = HttpResponse(out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="reporte_resumen.xlsx"'
        return resp


class PronosticoVentas(APIView):
    """
    Pronóstico de ventas para una fecha específica basado en datos históricos.
    Params: ?fecha=YYYY-MM-DD (opcional, default: mañana)
    Retorna productos que probablemente se venderán con estimación de cantidad.
    """
    def get(self, request):
        fecha_str = request.query_params.get('fecha')
        if fecha_str:
            try:
                fecha_objetivo = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            except:
                fecha_objetivo = timezone.now().date() + timedelta(days=1)
        else:
            fecha_objetivo = timezone.now().date() + timedelta(days=1)
        
        # Obtener día de la semana (0=lunes, 6=domingo)
        dia_semana = fecha_objetivo.weekday()
        
        # Obtener ventas históricas del mismo día de la semana en las últimas 8 semanas
        desde = fecha_objetivo - timedelta(days=56)  # 8 semanas atrás
        hasta = fecha_objetivo - timedelta(days=1)
        
        # Filtrar ventas del mismo día de la semana en el rango histórico
        ventas_historicas = Venta.objects.filter(
            fecha__date__gte=desde,
            fecha__date__lte=hasta
        ).select_related('cliente', 'sucursal')
        
        # Filtrar por día de la semana
        ventas_mismo_dia = []
        for venta in ventas_historicas:
            if venta.fecha.weekday() == dia_semana:
                ventas_mismo_dia.append(venta)
        
        # Calcular productos más vendidos en ese día de la semana
        productos_pronostico = {}
        
        for venta in ventas_mismo_dia:
            detalles = VentaDetalle.objects.filter(venta=venta).select_related('producto_variante__producto')
            for detalle in detalles:
                if detalle.producto_variante and detalle.producto_variante.producto:
                    producto_id = detalle.producto_variante.producto.id
                    producto_nombre = detalle.producto_variante.producto.nombre
                    
                    if producto_id not in productos_pronostico:
                        productos_pronostico[producto_id] = {
                            'producto_id': producto_id,
                            'producto_nombre': producto_nombre,
                            'cantidades': [],
                            'total_unidades': 0,
                            'veces_vendido': 0
                        }
                    
                    productos_pronostico[producto_id]['cantidades'].append(detalle.cantidad)
                    productos_pronostico[producto_id]['total_unidades'] += detalle.cantidad
                    productos_pronostico[producto_id]['veces_vendido'] += 1
        
        # Calcular promedios y estimaciones
        resultados = []
        for producto_id, datos in productos_pronostico.items():
            if len(datos['cantidades']) > 0:
                promedio = sum(datos['cantidades']) / len(datos['cantidades'])
                # Estimación: promedio redondeado hacia arriba
                estimacion = int(promedio) + (1 if promedio % 1 > 0.3 else 0)
                
                resultados.append({
                    'producto_id': producto_id,
                    'producto_nombre': datos['producto_nombre'],
                    'estimacion_unidades': estimacion,
                    'promedio_historico': round(promedio, 2),
                    'veces_vendido': datos['veces_vendido'],
                    'confianza': min(100, int((datos['veces_vendido'] / len(ventas_mismo_dia)) * 100)) if len(ventas_mismo_dia) > 0 else 0
                })
        
        # Ordenar por estimación descendente
        resultados.sort(key=lambda x: x['estimacion_unidades'], reverse=True)
        
        return Response({
            'fecha_pronostico': fecha_objetivo.strftime('%Y-%m-%d'),
            'dia_semana': ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][dia_semana],
            'productos': resultados[:20],  # Top 20
            'total_productos': len(resultados)
        })

